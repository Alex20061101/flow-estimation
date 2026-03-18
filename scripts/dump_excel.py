import openpyxl
import sys

def dump_sheet(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=False)
    
    with open("temp_dump.txt", "w", encoding="utf-8") as f:
        f.write(f"Sheets: {wb.sheetnames}\n")
        
        for idx in range(1, 3): # Dump second and third sheet just in case
            if idx >= len(wb.sheetnames): break
            sheet = wb[wb.sheetnames[idx]]
            f.write(f"\n--- Dumping sheet: {sheet.title} ---\n")
            
            for r in range(1, min(50, sheet.max_row + 1)):
                row_vals = []
                for c in range(1, min(60, sheet.max_column + 1)):
                    cell = sheet.cell(row=r, column=c)
                    if cell.value is not None:
                        row_vals.append(f"{openpyxl.utils.get_column_letter(c)}{r}: {cell.value}")
                if row_vals:
                    f.write(f"Row {r}: " + " | ".join(row_vals) + "\n")

if __name__ == "__main__":
    dump_sheet(sys.argv[1])
