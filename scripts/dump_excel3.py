import openpyxl

def find_formula_blocks(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=False)
    sheet = wb['Computing example for March']
    
    with open("temp_dump3.txt", "w", encoding="utf-8") as f:
        for r in range(1, 100):
            row_vals = []
            for c in range(1, 40):
                cell = sheet.cell(row=r, column=c)
                if isinstance(cell.value, str):
                    row_vals.append(f"{openpyxl.utils.get_column_letter(c)}{r}: {cell.value}")
            if row_vals:
                f.write(f"Row {r}: " + " | ".join(row_vals) + "\n")

if __name__ == "__main__":
    find_formula_blocks("c:\\Users\\alext\\Projects\\stf\\Both_March 1st.xlsx")
