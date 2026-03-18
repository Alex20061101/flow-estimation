import openpyxl

def dump_sheet(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=False)
    sheet = wb['Computing example for March']
    
    with open("temp_dump2.txt", "w", encoding="utf-8") as f:
        for r in range(1, min(50, sheet.max_row + 1)):
            row_vals = []
            for c in range(1, min(60, sheet.max_column + 1)):
                cell = sheet.cell(row=r, column=c)
                if cell.value is not None:
                    row_vals.append(f"{openpyxl.utils.get_column_letter(c)}{r}: {cell.value}")
            if row_vals:
                f.write(f"Row {r}: " + " | ".join(row_vals) + "\n")

if __name__ == "__main__":
    dump_sheet("c:\\Users\\alext\\Projects\\stf\\Both_March 1st.xlsx")
